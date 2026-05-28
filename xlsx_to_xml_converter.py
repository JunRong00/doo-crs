# python xlsx_to_xml_converter.py CRS_Template_Revised_2.xlsx --split 5 --out output_xml

#!/usr/bin/env python3
"""
FC XML Schema v2.2 Excel to XML Converter
Converts a CRS template Excel file to FC XML Schema v2.2 files for submission
to the Vanuatu MDES portal (https://mdes.doft.gov.vu/MDES/).

Reference: FC XML Schema v2.2 User Guide v1.0
           Vanuatu Customs and Inland Revenue

Usage:
    python xlsx_to_xml_converter.py <input.xlsx> [--split N] [--out <folder>]

Arguments:
    input.xlsx       Path to the filled CRS Excel template.
    --split N        Split into N XML files with balanced rows. Default: 1.
    --out <folder>   Output folder for XML files (created if missing).
                     Default: same folder as input.

Outputs:
    1. One or more FC XML v2.2 files.
    2. A copy of the Excel with a 'filepath' column on each account sheet.
"""

import sys
import os
import math
import uuid
import shutil
import argparse
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

# ── Namespaces ────────────────────────────────────────────────────────────────
# FC XML Schema v2.2 namespaces.
# Do NOT call ET.register_namespace() — Python reserves the ns\d+ prefix
# pattern for internal auto-generation and raises ValueError if you try.
# ElementTree auto-assigns ns0, ns1 in the order URIs are first encountered.
NS_MAIN  = "urn:fatcacrs:ties:v2"             # → auto-assigned ns0
NS_TYPES = "urn:oecd:ties:fatcacrstypes:v2"   # → auto-assigned ns1
NS_STF   = "urn:oecd:ties:stffatcatypes:v2"   # → auto-assigned ns2 (Address + Name children)

def ft(tag): return f"{{{NS_MAIN}}}{tag}"    # ns0: wrapper only (FATCA_CRS, MessageHeader, MessageBody)
def tp(tag): return f"{{{NS_TYPES}}}{tag}"   # ns1: main content elements
def sf(tag): return f"{{{NS_STF}}}{tag}"     # ns2: Address children + Name children (stffatcatypes)

# ── Template layout ────────────────────────────────────────────────────────────
REMARK_ROWS    = 6
DATA_START_ROW = 8

# ── Helpers ────────────────────────────────────────────────────────────────────

def safe(val):
    """Return stripped string or None for blank/NaN."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    s = str(val).strip()
    return s if s else None

def sub(parent, tag, text=None, attrib=None):
    el = ET.SubElement(parent, tag, attrib or {})
    if text is not None:
        el.text = text
    return el

def make_doc_ref(tc, sending_in):
    return f"{tc}{sending_in or ''}{uuid.uuid4()}"

def read_excluded_jurisdictions(xl_dict):
    key = next((k for k in xl_dict if k.strip().lower() == "excludedjurisdictions"), None)
    if key is None:
        return set()
    df = xl_dict[key]
    if len(df) < 7:
        return set()
    df = df.where(pd.notna(df), None)
    codes = set()
    for val in df.iloc[6:].iloc[:, 2]:
        code = safe(val)
        if code:
            codes.add(code.upper())
    return codes

def bool_attr(val):
    if val is None:
        return None
    return "true" if str(val).strip().lower() in ("true", "1", "yes") else "false"

def fmt_date(val):
    if not val:
        return None
    try:
        return pd.to_datetime(val).strftime("%Y-%m-%d")
    except Exception:
        return str(val).strip()

def fmt_amount(val):
    if val is None:
        return "0.00"
    try:
        return f"{float(val):.2f}"
    except Exception:
        return "0.00"

def read_sheet(xl_dict, name):
    """
    Read a sheet by name (case-insensitive).
    Supports two template layouts:
      Old: row 1 = field names, skip REMARK_ROWS rows.
      New: has a 'Field →' row, then REMARK_ROWS metadata rows, then data.
    Always drops col A (the label column) and lowercases column names.
    """
    key = next((k for k in xl_dict if k.strip().lower() == name.lower()), None)
    if key is None:
        return pd.DataFrame()
    df = xl_dict[key]

    field_row_idx = None
    for i in range(min(8, len(df))):
        val = df.iloc[i, 0] if df.shape[1] > 0 else None
        if val is not None and str(val).strip().lower().startswith("field"):
            field_row_idx = i
            break

    if field_row_idx is not None:
        col_names = [safe(v) or f"_col{j}" for j, v in enumerate(df.iloc[field_row_idx])]
        df = df.iloc[field_row_idx + 1 + REMARK_ROWS:].reset_index(drop=True)
        df.columns = col_names
    else:
        if len(df) <= REMARK_ROWS:
            return pd.DataFrame()
        df = df.iloc[REMARK_ROWS:].reset_index(drop=True)

    if df.empty:
        return pd.DataFrame()

    if df.shape[1] > 0:
        df = df.iloc[:, 1:]
    df.columns = [str(c).strip().lower() for c in df.columns]
    df = df.where(pd.notna(df), None)
    df = df.dropna(how="all").reset_index(drop=True)
    return df

def row_dict(row):
    return {str(k).strip().lower(): v for k, v in row.items()}

def split_rows(df, n):
    total = len(df)
    if total == 0 or n <= 1:
        return [df]
    chunk_size = math.ceil(total / n)
    return [df.iloc[i:i + chunk_size].reset_index(drop=True)
            for i in range(0, total, chunk_size)]

def write_xml(root, path):
    ET.indent(root, space="  ")
    with open(path, "w", encoding="utf-8") as f:
        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write(ET.tostring(root, encoding="unicode"))
        f.write("\n")

# ── XML element builders ───────────────────────────────────────────────────────

def build_doc_spec(parent, doc_type, doc_ref, corr_doc_ref=None):
    """DocSpec in FC types namespace (ns1). Values: OECD1/OECD2/OECD3."""
    ds = sub(parent, tp("DocSpec"))
    sub(ds, tp("DocTypeIndic"), doc_type or "OECD1")
    sub(ds, tp("DocRefId"),     doc_ref)
    if corr_doc_ref:
        sub(ds, tp("CorrDocRefId"), corr_doc_ref)
    return ds

def address_country(row):
    return (safe(row.get("address_countrycode")) or
            safe(row.get("countrycode")) or
            safe(row.get("rescountrycode")))

def build_res_country(parent, row, required=False, context="party"):
    country = (safe(row.get("rescountrycode")) or
               safe(row.get("address_countrycode")) or
               safe(row.get("countrycode")))
    if country:
        sub(parent, tp("ResCountryCode"), country)
        return country
    if required:
        raise ValueError(f"{context} is missing required ResCountryCode.")
    return None

def build_address(parent, row, required=False, context="party", force=False):
    """
    FC v2.2 Address.
    Attribute name: legalAddressType (per OECD XSD Address_Type).
    Address element: ns1 (tp). CountryCode, AddressFix and children: ns2 (cf/cfc).
    """
    country   = address_country(row)
    addr_type = safe(row.get("legaladdresstype"))  # column name unchanged in Excel
    addr_free = safe(row.get("addressfree"))
    street    = safe(row.get("street"))
    bldg      = safe(row.get("buildingidentifier"))
    suite     = safe(row.get("suiteidentifier"))
    floor_id  = safe(row.get("flooridentifier"))
    district  = safe(row.get("districtname"))
    pob       = safe(row.get("pob"))
    post_code = safe(row.get("postcode"))
    city      = safe(row.get("city"))
    subentity = safe(row.get("countrysubentity"))

    has_address = any([addr_free, street, bldg, suite, floor_id, district,
                       pob, post_code, city, subentity])
    if not country and required:
        raise ValueError(f"{context} is missing required address CountryCode.")
    if not country and not force and not has_address:
        return None

    attrib = {"legalAddressType": addr_type} if addr_type else {}   # per OECD XSD Address_Type
    addr_el = ET.SubElement(parent, tp("Address"), attrib)
    if country:
        sub(addr_el, sf("CountryCode"), country)   # CountryCode is in Address_Type (cfc/ns2)

    has_fix_fields = any([street, bldg, suite, floor_id, district,
                          pob, post_code, city, subentity])
    if has_fix_fields:
        fix = sub(addr_el, sf("AddressFix"))       # AddressFix and children are in cfc/ns2
        if street:    sub(fix, sf("Street"),             street)
        if bldg:      sub(fix, sf("BuildingIdentifier"), bldg)
        if suite:     sub(fix, sf("SuiteIdentifier"),    suite)
        if floor_id:  sub(fix, sf("FloorIdentifier"),    floor_id)
        if district:  sub(fix, sf("DistrictName"),       district)
        if pob:       sub(fix, sf("POB"),                pob)
        if post_code: sub(fix, sf("PostCode"),           post_code)
        if city:      sub(fix, sf("City"),               city)
        if subentity: sub(fix, sf("CountrySubentity"),   subentity)
        if addr_free:
            sub(addr_el, sf("AddressFree"), addr_free)
    elif addr_free:
        sub(addr_el, sf("AddressFree"), addr_free)
    else:
        sub(addr_el, sf("AddressFix"))
    return addr_el

def build_person_name(parent, row):
    """NamePerson_Type — nameType attribute (lowercase, per OECD XSD NamePerson_Type)."""
    name_type = safe(row.get("nametype"))
    attrib = {"nameType": name_type} if name_type else {}
    name_el = ET.SubElement(parent, tp("Name"), attrib)
    v = lambda k: safe(row.get(k))
    if v("precedingtitle"):       sub(name_el, sf("PrecedingTitle"),       v("precedingtitle"))
    if v("title"):                sub(name_el, sf("Title"),                v("title"))
    sub(name_el, sf("FirstName"), v("firstname") or "NFN")
    if v("middlename"):           sub(name_el, sf("MiddleName"),           v("middlename"))
    if v("nameprefix"):           sub(name_el, sf("NamePrefix"),           v("nameprefix"))
    sub(name_el, sf("LastName"),  v("lastname") or "UNKNOWN")
    if v("generationidentifier"): sub(name_el, sf("GenerationIdentifier"), v("generationidentifier"))
    if v("suffix"):               sub(name_el, sf("Suffix"),               v("suffix"))
    if v("generalsuffix"):        sub(name_el, sf("GeneralSuffix"),        v("generalsuffix"))
    return name_el

def build_org_name(parent, row):
    """Organisation Name — nameType attribute (lowercase, per OECD XSD NameOrganisation_Type)."""
    name_type = safe(row.get("nametype"))
    attrib = {"nameType": name_type} if name_type else {}
    el = ET.SubElement(parent, tp("Name"), attrib)
    el.text = safe(row.get("name")) or "UNKNOWN"
    return el

def build_tin(parent, row):
    """TIN — issuedBy attribute (lowercase, per OECD XSD TIN_Type)."""
    tin_val   = safe(row.get("tin"))
    issued_by = safe(row.get("tin_issuedby"))
    if not tin_val:
        return None
    attrib = {"issuedBy": issued_by} if issued_by else {}
    el = ET.SubElement(parent, tp("TIN"), attrib)
    el.text = tin_val
    return el

def build_org_in(parent, row):
    """Organisation TIN — FC v2.2 uses TIN (not IN) for organisations; issuedBy attribute only."""
    in_val    = safe(row.get("in"))
    issued_by = safe(row.get("in_issuedby"))
    if not in_val:
        return None
    attrib = {"issuedBy": issued_by} if issued_by else {}
    el = ET.SubElement(parent, tp("TIN"), attrib)
    el.text = in_val
    return el

def build_birth_info(parent, row):
    """BirthInfo — FC v2.2 main namespace."""
    birth_date = fmt_date(safe(row.get("birthdate")))
    birth_city = safe(row.get("birthcity"))
    birth_sub  = safe(row.get("birthcitysubentity"))
    birth_cc   = safe(row.get("birthcountrycode"))
    former_cc  = safe(row.get("formercountryname"))

    if not any([birth_date, birth_city, birth_cc, former_cc]):
        return None
    el = sub(parent, tp("BirthInfo"))
    if birth_date: sub(el, tp("BirthDate"),     birth_date)
    if birth_city: sub(el, tp("City"),           birth_city)
    if birth_sub:  sub(el, tp("CitySubentity"),  birth_sub)
    if birth_cc or former_cc:
        ci = sub(el, tp("CountryInfo"))
        if birth_cc:    sub(ci, tp("CountryCode"),       birth_cc)
        elif former_cc: sub(ci, tp("FormerCountryName"), former_cc)
    return el

def build_account_number(parent, row):
    """
    AccountNumber — FC v2.2.
    Attribute name: AcctNumberType (per OECD XSD FIAccountNumber_Type).
    """
    acc_num   = safe(row.get("accountnumber")) or "NANUM"
    acct_type = safe(row.get("acctnumbertype"))
    undoc     = bool_attr(safe(row.get("undocumentedaccount")))
    closed    = bool_attr(safe(row.get("closedaccount")))
    dormant   = bool_attr(safe(row.get("dormantaccount")))

    attrib = {}
    if undoc   is not None: attrib["UndocumentedAccount"] = undoc
    if closed  is not None: attrib["ClosedAccount"]       = closed
    if dormant is not None: attrib["DormantAccount"]      = dormant

    el = ET.SubElement(parent, tp("AccountNumber"), attrib)
    el.text = acc_num
    return el

def build_account_balance(parent, row):
    """AccountBalance — currCode attribute (lowercase, per OECD XSD MonAmnt_Type)."""
    curr = safe(row.get("currcode")) or "XXX"
    el   = ET.SubElement(parent, tp("AccountBalance"), {"currCode": curr})
    el.text = fmt_amount(safe(row.get("accountbalance")))
    return el

def build_payment(parent, pmt_row):
    """
    Payment — Type element + currCode attribute (lowercase), per OECD XSD Payment_Type.
    """
    pmt_type = safe(pmt_row.get("type"))
    if not pmt_type:
        return None
    curr    = safe(pmt_row.get("currcode")) or "XXX"
    pmt_el  = sub(parent, tp("Payment"))
    sub(pmt_el, tp("Type"), pmt_type)                                   # per XSD: element named Type
    amnt_el = ET.SubElement(pmt_el, tp("PaymentAmnt"), {"currCode": curr})  # per XSD: attr currCode
    amnt_el.text = fmt_amount(safe(pmt_row.get("paymentamnt")))
    return pmt_el

def build_controlling_person(parent, cp_row):
    """
    ControllingPerson — FC v2.2.
    Order: Individual → CtrlgPersonType.
    SelfCert removed. CtrlgPersonType values: CRS801–CRS813 (was CRS800).
    """
    cp_el  = sub(parent, tp("ControllingPerson"))
    ind_el = sub(cp_el,  tp("Individual"))

    acc = safe(cp_row.get("accountnumber")) or "unknown account"
    build_res_country(ind_el, cp_row, required=False,
                      context=f"ControllingPerson for {acc}")
    build_tin(ind_el, cp_row)
    build_person_name(ind_el, cp_row)
    build_address(ind_el, cp_row, required=True,
                  context=f"ControllingPerson for {acc}")
    build_birth_info(ind_el, cp_row)

    sub(cp_el, tp("CtrlgPersonType"), safe(cp_row.get("ctrlgpersontype")) or "CRS801")
    return cp_el

def cp_has_data(cp_row):
    return any(safe(cp_row.get(k)) for k in
               ("rescountrycode", "firstname", "lastname",
                "countrycode", "address_countrycode", "city", "addressfree"))

def build_account_index(df):
    idx = {}
    if df.empty or "accountnumber" not in df.columns:
        return idx
    for i, row in df.iterrows():
        key = safe(str(row.get("accountnumber")))
        if key:
            idx.setdefault(key, []).append((i, row_dict(row)))
    return idx

def get_linked(df, acc_num):
    if df.empty or not acc_num:
        return pd.DataFrame()
    return df[df["accountnumber"].astype(str).str.strip() == acc_num]

# ── Section builders ───────────────────────────────────────────────────────────

def build_message_header(root, header_row, tc, part_num=1, total_parts=1):
    """
    MessageHeader — FC v2.2 (replaces MessageSpec).
    MessageType is always "FATCA-CRS". CorrMessageRefId added (optional).
    Element order: SendingCompanyIN, TransmittingCountry, ReceivingCountry,
    MessageType, Warning, Contact, MessageRefId, MessageTypeIndic,
    CorrMessageRefId, ReportingPeriod, Timestamp.
    """
    hdr = sub(root, ft("MessageHeader"))

    sending_in = safe(header_row.get("sendingcompanyin"))
    if sending_in:
        sub(hdr, tp("SendingCompanyIN"), sending_in)

    rc = safe(header_row.get("receivingcountry")) or "XX"
    sub(hdr, tp("TransmittingCountry"), tc)
    sub(hdr, tp("ReceivingCountry"),    rc)
    sub(hdr, tp("MessageType"),         "FATCA-CRS")

    warning = safe(header_row.get("warning")) or ""
    if total_parts > 1:
        warning = f"{warning} Split file {part_num} of {total_parts}.".strip()
    if warning:
        sub(hdr, tp("Warning"), warning)

    contact = safe(header_row.get("contact"))
    if contact:
        sub(hdr, tp("Contact"), contact)

    rep_period = safe(header_row.get("reportingperiod")) or datetime.now().strftime("%Y-%m-%d")

    provided = safe(header_row.get("messagerefid"))
    if provided and total_parts == 1:
        msg_ref_id = provided
    else:
        suffix = f"P{part_num:02d}" if total_parts > 1 else ""
        msg_ref_id = f"{tc}2025{sending_in or ''}{uuid.uuid4().hex[:12].upper()}{suffix}"

    sub(hdr, tp("MessageRefId"),     msg_ref_id)
    sub(hdr, tp("MessageTypeIndic"), safe(header_row.get("messagetypeindic")) or "CRS701")

    corr_msg_ref = safe(header_row.get("corrmessagerefid"))
    if corr_msg_ref:
        sub(hdr, tp("CorrMessageRefId"), corr_msg_ref)

    sub(hdr, tp("ReportingPeriod"), fmt_date(rep_period) or rep_period)
    timestamp = safe(header_row.get("timestamp"))
    if not timestamp:
        timestamp = datetime.now().isoformat(timespec="seconds")
    sub(hdr, tp("Timestamp"), timestamp)
    return hdr

def build_reporting_fi(body, fi_row, tc, sending_in=None):
    """
    ReportingFI — FC v2.2.
    Goes directly inside MessageBody (no CrsBody wrapper).
    Order: ResCountryCode → IN → Name → Address → DocSpec.
    """
    fi_el = sub(body, tp("ReportingFI"))
    build_res_country(fi_el, fi_row, required=False)
    build_org_in(fi_el, fi_row)
    build_org_name(fi_el, fi_row)
    build_address(fi_el, fi_row, required=True, context="ReportingFI")
    build_doc_spec(fi_el,
                   safe(fi_row.get("doctypeindic"))  or "OECD1",
                   safe(fi_row.get("docrefid"))      or make_doc_ref(tc, sending_in),
                   safe(fi_row.get("corrdocrefid")))
    return fi_el

def build_individual_account(rg_el, rd, cp_idx, pmt_idx, tc, sending_in=None):
    """
    Individual AccountReport — FC v2.2.
    Removed: EquityInterestType, SelfCert, DDProcedure, AccountType, JointAccount.
    AccountHolder contains only Individual (no AcctHolderType for individuals).
    ControllingPerson placed after AccountHolder, before AccountBalance.
    """
    ar_el = sub(rg_el, tp("AccountReport"))
    build_doc_spec(ar_el,
                   safe(rd.get("doctypeindic"))  or "OECD1",
                   safe(rd.get("docrefid"))      or make_doc_ref(tc, sending_in),
                   safe(rd.get("corrdocrefid")))
    build_account_number(ar_el, rd)

    ah_el  = sub(ar_el, tp("AccountHolder"))
    ind_el = sub(ah_el, tp("Individual"))
    build_res_country(ind_el, rd, required=True,
                      context=f"Individual account {safe(rd.get('accountnumber')) or 'unknown'}")
    build_tin(ind_el, rd)
    build_person_name(ind_el, rd)
    build_address(ind_el, rd, required=True,
                  context=f"Individual account {safe(rd.get('accountnumber')) or 'unknown'}")
    build_birth_info(ind_el, rd)

    acc_num = safe(rd.get("accountnumber"))
    for _, cp_row in cp_idx.get(acc_num, []):
        if cp_has_data(cp_row):
            build_controlling_person(ar_el, cp_row)

    build_account_balance(ar_el, rd)

    for _, pmt_row in pmt_idx.get(acc_num, []):
        build_payment(ar_el, pmt_row)

def build_organisation_account(rg_el, rd, cp_idx, pmt_idx, tc, sending_in=None):
    """
    Organisation AccountReport — FC v2.2.
    Removed: EquityInterestType, SelfCert, DDProcedure, AccountType, JointAccount.
    AcctHolderType → AcctHolderTypeCRS (sibling of Organisation in AccountHolder).
    """
    ar_el = sub(rg_el, tp("AccountReport"))
    build_doc_spec(ar_el,
                   safe(rd.get("doctypeindic"))  or "OECD1",
                   safe(rd.get("docrefid"))      or make_doc_ref(tc, sending_in),
                   safe(rd.get("corrdocrefid")))
    build_account_number(ar_el, rd)

    ah_el  = sub(ar_el, tp("AccountHolder"))
    org_el = sub(ah_el, tp("Organisation"))
    build_res_country(org_el, rd, required=False)
    build_org_in(org_el, rd)
    build_org_name(org_el, rd)
    build_address(org_el, rd, required=False,
                  context=f"Organisation account {safe(rd.get('accountnumber')) or 'unknown'}",
                  force=True)

    acct_holder_type = safe(rd.get("acctholdertype"))
    if acct_holder_type:
        sub(ah_el, tp("AcctHolderTypeCRS"), acct_holder_type)

    acc_num = safe(rd.get("accountnumber"))
    for _, cp_row in cp_idx.get(acc_num, []):
        if cp_has_data(cp_row):
            build_controlling_person(ar_el, cp_row)

    build_account_balance(ar_el, rd)

    for _, pmt_row in pmt_idx.get(acc_num, []):
        build_payment(ar_el, pmt_row)

# ── Excel filepath column writer ───────────────────────────────────────────────

def write_filepath_column(input_path, output_excel_path,
                          ind_map, org_map, cp_map, pmt_map):
    shutil.copy2(input_path, output_excel_path)
    wb = openpyxl.load_workbook(output_excel_path)

    fp_fill  = PatternFill("solid", fgColor="FFF2CC")
    fp_font  = Font(name="Arial", size=10, italic=True, color="7F6000")
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    sheet_maps = {
        "Individual":        ind_map,
        "Organisation":      org_map,
        "ControllingPerson": cp_map,
        "Payment":           pmt_map,
    }

    for sheet_name, fp_map in sheet_maps.items():
        if sheet_name not in wb.sheetnames or not fp_map:
            continue
        ws      = wb[sheet_name]
        new_col = ws.max_column + 1
        col_ltr = get_column_letter(new_col)

        field_row = 1
        for row_num in range(1, min(ws.max_row, 10) + 1):
            val = ws.cell(row=row_num, column=1).value
            if val is not None and str(val).strip().lower().startswith("field"):
                field_row = row_num
                break
        data_start_row = field_row + 1 + REMARK_ROWS

        hdr = ws.cell(row=field_row, column=new_col, value="filepath")
        hdr.fill = hdr_fill
        hdr.font = hdr_font

        for rr in range(field_row + 1, data_start_row):
            ws.cell(row=rr, column=new_col, value="")

        for data_idx, filepath in fp_map.items():
            excel_row = data_start_row + data_idx
            c = ws.cell(row=excel_row, column=new_col, value=filepath)
            c.fill = fp_fill
            c.font = fp_font

        ws.column_dimensions[col_ltr].width = 45

    wb.save(output_excel_path)

# ── Main converter ─────────────────────────────────────────────────────────────

def convert(input_path, n_splits=1, output_prefix=None):
    print(f"\n{'='*60}")
    print(f"  FC XML Schema v2.2 Converter")
    print(f"{'='*60}")
    print(f"  Input  : {input_path}")
    print(f"  Splits : {n_splits}")

    xl = pd.read_excel(input_path, sheet_name=None, dtype=str, header=0, keep_default_na=False)

    header_df = read_sheet(xl, "MessageHeader")
    fi_df     = read_sheet(xl, "ReportingFI")
    ind_df    = read_sheet(xl, "Individual")
    org_df    = read_sheet(xl, "Organisation")
    cp_df     = read_sheet(xl, "ControllingPerson")
    pmt_df    = read_sheet(xl, "Payment")

    if header_df.empty:
        raise ValueError("MessageHeader sheet is missing or has no data rows.")

    header_row = row_dict(header_df.iloc[0])
    tc         = safe(header_row.get("transmittingcountry")) or "MY"
    mti        = safe(header_row.get("messagetypeindic"))    or "CRS701"
    sending_in = safe(header_row.get("sendingcompanyin"))

    base    = Path(input_path).stem
    in_dir  = Path(input_path).parent
    if output_prefix is None:
        out_dir = in_dir
    elif "/" in output_prefix or output_prefix.endswith(os.sep):
        out_dir = Path(output_prefix)
        out_dir.mkdir(parents=True, exist_ok=True)
    else:
        out_dir = in_dir / output_prefix
        out_dir.mkdir(parents=True, exist_ok=True)

    def xml_path(part, total):
        if total == 1:
            return str(out_dir / f"{base}_CRS.xml")
        return str(out_dir / f"{base}_CRS_part{part:02d}.xml")

    excel_out = str(in_dir / f"{base}_with_filepath.xlsx")

    # ── Nil return ─────────────────────────────────────────────────────────────
    if mti == "CRS703":
        print("  Mode   : Nil return (CRS703)")
        if fi_df.empty:
            raise ValueError("ReportingFI sheet is missing or has no data rows.")
        fi_row = row_dict(fi_df.iloc[0])
        root   = ET.Element(ft("FATCA_CRS"), {"version": "2.2"})
        build_message_header(root, header_row, tc)
        body   = sub(root, ft("MessageBody"))
        build_reporting_fi(body, fi_row, tc, sending_in)
        rg_el  = sub(body, tp("ReportingGroup"))
        nil_el = sub(rg_el, tp("NilReport"))
        build_doc_spec(nil_el,
                       safe(fi_row.get("doctypeindic")) or "OECD1",
                       safe(fi_row.get("docrefid")) or make_doc_ref(tc, sending_in))
        op = xml_path(1, 1)
        write_xml(root, op)
        print(f"\n  ✅  {op}")
        return [op], None

    if fi_df.empty:
        raise ValueError("ReportingFI sheet is missing or has no data rows.")
    fi_row = row_dict(fi_df.iloc[0])

    # ── Filter excluded jurisdictions ──────────────────────────────────────────
    excluded = read_excluded_jurisdictions(xl)
    if excluded and not ind_df.empty and "rescountrycode" in ind_df.columns:
        before = len(ind_df)
        ind_df = ind_df[~ind_df["rescountrycode"].str.upper().isin(excluded)].reset_index(drop=True)
        if len(ind_df) < before:
            print(f"  Filtered: {before - len(ind_df)} Individual row(s) excluded")
    if excluded and not org_df.empty and "rescountrycode" in org_df.columns:
        before = len(org_df)
        org_df = org_df[~org_df["rescountrycode"].str.upper().isin(excluded)].reset_index(drop=True)
        if len(org_df) < before:
            print(f"  Filtered: {before - len(org_df)} Organisation row(s) excluded")

    # ── Tag account rows ───────────────────────────────────────────────────────
    ind_tagged = ind_df.copy()
    ind_tagged["__sheet__"]    = "Individual"
    ind_tagged["__orig_idx__"] = range(len(ind_tagged))

    org_tagged = org_df.copy()
    org_tagged["__sheet__"]    = "Organisation"
    org_tagged["__orig_idx__"] = range(len(org_tagged))

    all_accounts   = pd.concat([ind_tagged, org_tagged], ignore_index=True)
    total_accounts = len(all_accounts)

    if total_accounts == 0:
        raise ValueError("No account rows found in Individual or Organisation sheets.")

    actual_splits = min(n_splits, total_accounts)
    if actual_splits < n_splits:
        print(f"  ⚠️   Only {total_accounts} account(s) — reducing to {actual_splits} file(s).")

    chunks      = split_rows(all_accounts, actual_splits)
    total_parts = len(chunks)

    print(f"  Accounts: {total_accounts} rows → {total_parts} file(s)")
    print(f"{'='*60}\n")

    cp_idx  = build_account_index(cp_df)
    pmt_idx = build_account_index(pmt_df)

    ind_map = {}
    org_map = {}
    cp_map  = {}
    pmt_map = {}
    xml_files = []

    # ── One XML per chunk ──────────────────────────────────────────────────────
    for part_num, chunk in enumerate(chunks, start=1):
        op   = xml_path(part_num, total_parts)
        root = ET.Element(ft("FATCA_CRS"), {"version": "2.2"})

        build_message_header(root, header_row, tc, part_num, total_parts)

        body  = sub(root, ft("MessageBody"))
        build_reporting_fi(body, fi_row, tc, sending_in)
        rg_el = sub(body, tp("ReportingGroup"))

        for _, row in chunk.iterrows():
            rd       = row_dict(row)
            sheet    = rd.pop("__sheet__",    None)
            orig_idx = int(rd.pop("__orig_idx__", -1))
            acc_num  = safe(rd.get("accountnumber"))

            if not acc_num:
                continue

            if sheet == "Individual":
                build_individual_account(rg_el, rd, cp_idx, pmt_idx, tc, sending_in)
                ind_map[orig_idx] = op
            elif sheet == "Organisation":
                build_organisation_account(rg_el, rd, cp_idx, pmt_idx, tc, sending_in)
                org_map[orig_idx] = op

            for ci, _ in cp_idx.get(acc_num, []):
                cp_map[ci] = op
            for pi, _ in pmt_idx.get(acc_num, []):
                pmt_map[pi] = op

        write_xml(root, op)
        print(f"  Part {part_num:02d}/{total_parts:02d} "
              f"— {len(chunk):>4} account row(s) → {Path(op).name}")
        xml_files.append(op)

    # ── Write Excel with filepath column ───────────────────────────────────────
    write_filepath_column(input_path, excel_out,
                          ind_map, org_map, cp_map, pmt_map)

    print(f"\n  📊  Excel with filepath → {Path(excel_out).name}")
    print(f"\n{'='*60}")
    print(f"  Done. {total_parts} XML file(s) + 1 Excel file generated.")
    print(f"{'='*60}\n")

    return xml_files, excel_out

# ── CLI ────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="FC XML Schema v2.2 Excel to XML Converter",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument("input",
                        help="Input Excel file (.xlsx)")
    parser.add_argument("--split", "-s", type=int, default=1,
                        metavar="N",
                        help="Number of XML files to split into (default: 1)")
    parser.add_argument("--out", "-o", default=None,
                        metavar="FOLDER",
                        help="Output folder for XML files (created if missing).")
    args = parser.parse_args()

    if args.split < 1:
        print("Error: --split must be >= 1")
        sys.exit(1)

    convert(args.input, n_splits=args.split, output_prefix=args.out)
